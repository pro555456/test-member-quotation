import argparse
import json
import os
import re
import sys
from datetime import datetime, date

import openpyxl
from openpyxl.utils.datetime import from_excel


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def normalize_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.strftime("%Y-%m-%d")
            if isinstance(converted, date):
                return converted.strftime("%Y-%m-%d")
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return None

    cleaned = (
        text.replace(".", "/")
        .replace("年", "/")
        .replace("月", "/")
        .replace("日", "")
        .replace("上午", "")
        .replace("下午", "")
        .replace("AM", "")
        .replace("PM", "")
        .replace("am", "")
        .replace("pm", "")
        .strip()
    )

    match = re.match(r"^(\d{2,4})[/-](\d{1,2})[/-](\d{1,2})", cleaned)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))
        if year < 1000:
            year += 1911
        try:
            return date(year, month, day).strftime("%Y-%m-%d")
        except ValueError:
            return cleaned[:10]

    return cleaned[:10]


def normalize_number(value, default=0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:
        return default


def normalize_int(value, default=0):
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except Exception:
        return default


def normalize_bool(value):
    return 1 if normalize_int(value, 0) > 0 else 0


def parse_rows(workbook_path, sheet_name):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    rows = []

    for row_no in range(4, ws.max_row + 1):
        seq = ws.cell(row_no, 1).value
        quote_date = normalize_date(ws.cell(row_no, 2).value)
        customer_order_no = ws.cell(row_no, 3).value
        customer_name = ws.cell(row_no, 4).value
        game_title = ws.cell(row_no, 5).value

        if seq in (None, "") and not quote_date and not customer_name and not game_title:
            continue
        if not customer_name or not game_title:
            continue

        rows.append({
            "sourceRowNo": row_no,
            "quoteDate": quote_date,
            "customerOrderNo": str(customer_order_no).strip() if customer_order_no not in (None, "") else None,
            "customerName": str(customer_name).strip(),
            "gameTitle": str(game_title).strip(),
            "platforms": {
                "ios": normalize_bool(ws.cell(row_no, 6).value),
                "android": normalize_bool(ws.cell(row_no, 7).value),
                "web": normalize_bool(ws.cell(row_no, 8).value),
                "other": normalize_bool(ws.cell(row_no, 9).value),
            },
            "signedAt": normalize_date(ws.cell(row_no, 10).value),
            "notes": str(ws.cell(row_no, 11).value).strip() if ws.cell(row_no, 11).value not in (None, "") else None,
            "internalOrderNo": str(ws.cell(row_no, 13).value).strip() if ws.cell(row_no, 13).value not in (None, "") else None,
            "quantity": normalize_int(ws.cell(row_no, 14).value, 1) or 1,
            "unitPriceUntaxed": normalize_number(ws.cell(row_no, 15).value, 0),
            "totalUntaxed": normalize_number(ws.cell(row_no, 16).value, 0),
            "closedAt": normalize_date(ws.cell(row_no, 17).value),
        })

    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sheet", default=os.environ.get("QUOTE_IMPORT_DEFAULT_SHEET", "2025"))
    args = parser.parse_args()

    workbook_path = os.environ.get("QUOTE_XLSX_PATH")
    if not workbook_path:
        raise RuntimeError("QUOTE_XLSX_PATH is required")

    rows = parse_rows(workbook_path, args.sheet)
    json.dump({"sheet": args.sheet, "rows": rows}, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
